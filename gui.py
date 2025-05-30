    from tkinter import (
        Tk,
        Canvas,
        Label,
        Entry,
        Button,
        messagebox,
        StringVar,
        Toplevel,
        Frame,
        LEFT,
        SOLID,
        X,
        RIGHT,
        Listbox,
    )
    import psycopg2
    from hashlib import sha256
    import datetime
    import re
    import os

    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        openpyxl = None

    # Конфигурация подключения к БД
    DB_CONFIG = {
        "dbname": "postgres",
        "user": "postgres",
        "password": "",
        "host": "127.0.0.1",
        "port": "5432",
        "options": "-c search_path=Библиотека",
    }

    # Глобальные переменные
    PASSWORD_CACHE = {}
    bg_color = "#B18653"
    fg_color = "#C3A580"
    text_color = "#2A1E17"

    # Создание основного окна
    root = Tk()
    root.state("zoomed")
    width = root.winfo_screenwidth()
    height = root.winfo_screenheight()

    # Создание Canvas
    canvas = Canvas(root, bg=bg_color)
    canvas.pack(expand=True, fill="both")


    def clear_window():
        """Очищает все элементы интерфейса"""
        for widget in root.winfo_children():
            if widget != canvas:  # Не удаляем canvas
                widget.destroy()
        canvas.delete("all")


    def round_rectangle(x1, y1, x2, y2, radius=25, **kwargs):
        """Создает скругленный прямоугольник"""
        points = [
            x1 + radius,
            y1,
            x2 - radius,
            y1,
            x2,
            y1,
            x2,
            y1 + radius,
            x2,
            y2 - radius,
            x2,
            y2,
            x2 - radius,
            y2,
            x1 + radius,
            y2,
            x1,
            y2,
            x1,
            y2 - radius,
            x1,
            y1 + radius,
            x1,
            y1,
            x1 + radius,
            y1,
        ]
        return canvas.create_polygon(points, **kwargs, smooth=True)


    def create_login_window():
        """Создает окно входа"""
        clear_window()

        # Создание фонового прямоугольника
        rect_width = 400
        rect_height = 300
        x1 = (width - rect_width) // 2
        y1 = (height - rect_height) // 2
        x2 = x1 + rect_width
        y2 = y1 + rect_height

        round_rectangle(x1, y1, x2, y2, radius=20, fill=fg_color, stipple="gray25")

        # Элементы интерфейса входа
        Label(root, text="Вход в библиотеку", font=("Arial", 16), bg=fg_color, fg=text_color).place(
            x=width // 2 - 75, y=y1 + 30
        )

        Label(root, text="ФИО:", bg=fg_color, fg=text_color).place(x=x1 + 50, y=y1 + 80)
        login_entry = Entry(root, width=40)
        login_entry.place(x=x1 + 110, y=y1 + 80)

        Label(root, text="Пароль:", bg=fg_color, fg=text_color).place(x=x1 + 50, y=y1 + 120)
        password_entry = Entry(root, width=40, show="*")
        password_entry.place(x=x1 + 110, y=y1 + 120)

        def check_credentials():
            login = login_entry.get()
            password = password_entry.get()

            if not login or not password:
                messagebox.showerror("Ошибка", "Заполните все поля")
                return

            try:
                conn = psycopg2.connect(**DB_CONFIG)
                cursor = conn.cursor()

                # Сначала проверяем, является ли пользователь сотрудником
                cursor.execute('SELECT "Пароль" FROM "Сотрудник" WHERE "ФИО" = %s', (login,))
                staff_result = cursor.fetchone()

                if staff_result:
                    if str(staff_result[0]) == password:
                        PASSWORD_CACHE[login] = sha256(password.encode()).hexdigest()
                        create_staff_home(login)  # Переходим на главную страницу сотрудника
                        return
                    else:
                        messagebox.showerror("Ошибка", "Неверный пароль")
                        return

                # Если не сотрудник, проверяем как читателя
                cursor.execute('SELECT "Пароль" FROM "Читатели" WHERE "ФИО" = %s', (login,))
                reader_result = cursor.fetchone()

                if reader_result:
                    if str(reader_result[0]) == password:
                        PASSWORD_CACHE[login] = sha256(password.encode()).hexdigest()
                        create_user_home(login)  # Переходим на главную страницу читателя
                    else:
                        messagebox.showerror("Ошибка", "Неверный пароль")
                else:
                    messagebox.showerror("Ошибка", "Пользователь не найден")

            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка подключения к БД: {str(e)}")
            finally:
                if "conn" in locals():
                    conn.close()

        Button(root, text="Вход", command=check_credentials, bg="#8B7355", fg="white").place(x=width // 2 - 25, y=y1 + 170)

        Label(root, text="Забыли пароль?", fg=text_color, bg=fg_color, cursor="hand2").place(x=width // 2 - 50, y=y1 + 210)

        # Кнопка перехода к регистрации
        Button(root, text="Регистрация", command=create_registration_window, bg="#8B7355", fg="white").place(
            x=width - 120, y=20
        )


    def create_registration_window():
        """Создает окно регистрации"""
        clear_window()

        # Создание фонового прямоугольника
        rect_width = 500
        rect_height = 400
        x1 = (width - rect_width) // 2
        y1 = (height - rect_height) // 2
        x2 = x1 + rect_width
        y2 = y1 + rect_height

        round_rectangle(x1, y1, x2, y2, radius=20, fill=fg_color, stipple="gray25")

        # Элементы интерфейса регистрации
        Label(root, text="Регистрация", font=("Arial", 16), bg=fg_color, fg=text_color).place(x=width // 2 - 50, y=y1 + 30)

        # Поля для ввода данных
        fields = [
            ("ФИО", y1 + 80),
            ("Номер телефона", y1 + 120),
            ("Адрес", y1 + 160),
            ("Пароль", y1 + 200),
            ("Повторите пароль", y1 + 240),
        ]

        entries = []
        for text, y_pos in fields:
            Label(root, text=text + ":", bg=fg_color, fg=text_color).place(x=x1 + 50, y=y_pos)
            entry = Entry(root, width=40)
            if "Пароль" in text:
                entry.config(show="*")
            entry.place(x=x1 + 180, y=y_pos)
            entries.append(entry)

        def register_user():
            # Получаем данные из полей ввода
            data = [entry.get() for entry in entries]
            if not all(data):
                messagebox.showerror("Ошибка", "Все поля должны быть заполнены")
                return

            if data[3] != data[4]:
                messagebox.showerror("Ошибка", "Пароли не совпадают")
                return

            try:
                conn = psycopg2.connect(**DB_CONFIG)
                cursor = conn.cursor()

                # Проверяем, есть ли уже такой пользователь
                cursor.execute('SELECT "ФИО" FROM "Читатели" WHERE "ФИО" = %s', (data[0],))
                if cursor.fetchone():
                    messagebox.showerror("Ошибка", "Пользователь с таким ФИО уже существует")
                    return

                # Добавляем нового пользователя
                cursor.execute(
                    """INSERT INTO "Читатели" ("ФИО", "Номер телефона", "Адрес", "Пароль") 
                    VALUES (%s, %s, %s, %s)""",
                    (data[0], data[1], data[2], data[3]),
                )
                conn.commit()
                messagebox.showinfo("Успех", "Регистрация прошла успешно!")
                create_login_window()

            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка при регистрации: {str(e)}")
            finally:
                if "conn" in locals():
                    conn.close()

        Button(root, text="Регистрация", command=register_user, bg="#8B7355", fg="white").place(
            x=width // 2 - 40, y=y1 + 300
        )

        # Кнопка перехода ко входу
        Button(root, text="Вход", command=logout, bg="#8B7355", fg="white").place(x=width - 100, y=20)


    # Создаем первоначальное окно входа
    create_login_window()


    def create_user_home(user_fio, offset=0):
        """Создает главную страницу пользователя с новым дизайном"""
        clear_window()

        # Устанавливаем новый цвет фона
        canvas.config(bg="#D5C5AE")

        # Создаем верхнюю полосу с ФИО пользователя
        header_height = 60
        canvas.create_rectangle(
            0, 0, width, header_height, fill="#54101D", stipple="gray50", outline=""  # Полупрозрачность
        )

        # Добавляем ФИО пользователя по центру полосы
        Label(root, text=user_fio, font=("Arial", 16), bg="#54101D", fg="white").place(
            x=width // 2, y=header_height // 2, anchor="center"
        )

        # Создаем круг с инициалами в правом верхнем углу (сделаем его кнопкой)
        initials = "".join([name[0] for name in user_fio.split()[:2]])
        circle_radius = 25
        circle_x = width - 40
        circle_y = header_height // 2

        # Создаем кнопку с инициалами
        initials_btn = Button(
            root,
            text=initials,
            font=("Arial", 12),
            bg="white",
            fg="#54101D",
            bd=0,
            command=lambda: create_personal_account(user_fio),
        )
        initials_btn.place(
            x=circle_x - circle_radius, y=circle_y - circle_radius, width=circle_radius * 2, height=circle_radius * 2
        )

        Label(root, text=initials, font=("Arial", 12), bg="white", fg="#54101D").place(
            x=circle_x, y=circle_y, anchor="center"
        )

        # Получаем данные о выданных книгах из БД
        try:
            conn = psycopg2.connect(**DB_CONFIG)
            cursor = conn.cursor()

            # 1. Находим ID читателя по ФИО
            cursor.execute('SELECT "id" FROM "Читатели" WHERE "ФИО" = %s', (user_fio,))
            reader_id = cursor.fetchone()[0]

            # 2. Получаем список выданных книг для этого читателя (только не возвращенные)
            cursor.execute(
                """
                SELECT k."Название", k."Автор", sv."Количество", 
                    v."Дата выдачи", v."Дата возврата"
                FROM "Выдача" v
                JOIN "Состав выдачи" sv ON v."id" = sv."Выдача"
                JOIN "Книги" k ON sv."Книга" = k."id"
                WHERE v."Читатель" = %s AND v."Создана" = FALSE
                ORDER BY v."Дата возврата" ASC
            """,
                (reader_id,),
            )

            all_books = cursor.fetchall()
            total_books = len(all_books)

            # Вычисляем корректный offset для циклической прокрутки
            if offset >= total_books:
                offset = 0
            elif offset < 0:
                offset = total_books - (total_books % 3 or 3)

            # Берем только 3 книги, начиная с текущего offset
            books = all_books[offset : offset + 3]

            # Отображаем книги в центре экрана
            book_y = header_height + 50
            book_width = 600
            book_height = 150

            for i, (title, author, quantity, issue_date, return_date) in enumerate(books):
                # Создаем полупрозрачный белый блок для книги
                canvas.create_rectangle(
                    (width - book_width) // 2,
                    book_y,
                    (width + book_width) // 2,
                    book_y + book_height,
                    fill="white",
                    stipple="gray75",  # Полупрозрачность
                    outline="",
                )

                title_label = Label(
                    root,
                    text=title,
                    font=("Arial", 14),
                    bg="white",
                    fg=text_color,
                    width=30,
                    wraplength=book_width - 40,
                    anchor="center",
                )
                title_label.place(x=width // 2, y=book_y + 20, anchor="center")

                Label(root, text=f"Автор: {author}", bg="white", fg=text_color).place(
                    x=width // 2 - book_width // 2 + 20, y=book_y + 50
                )

                Label(root, text=f"Количество: {quantity}", bg="white", fg=text_color).place(
                    x=width // 2 - book_width // 2 + 20, y=book_y + 80
                )

                Label(root, text=f"Дата выдачи: {issue_date.strftime('%d.%m.%Y')}", bg="white", fg=text_color).place(
                    x=width // 2 + book_width // 2 - 20, y=book_y + 50, anchor="e"
                )

                Label(root, text=f"Дата возврата: {return_date.strftime('%d.%m.%Y')}", bg="white", fg=text_color).place(
                    x=width // 2 + book_width // 2 - 20, y=book_y + 80, anchor="e"
                )

                book_y += book_height + 30

            # Если книг нет
            if not all_books:
                Label(root, text="У вас нет взятых книг", font=("Arial", 14), bg="#D5C5AE", fg=text_color).place(
                    x=width // 2, y=height // 2, anchor="center"
                )

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка получения данных: {str(e)}")
        finally:
            if "conn" in locals():
                conn.close()

            # Всегда показываем кнопку "Далее" (если есть книги)
        if total_books > 3:
            Button(
                root,
                text="Далее",
                bg="#946A66",
                fg="white",
                activebackground="#946A66",
                activeforeground="white",
                relief="flat",
                command=lambda: create_user_home(user_fio, offset + 3),
            ).place(x=100, y=700)

            # Всегда показываем кнопку "Назад" (если есть книги)
        if total_books > 3:
            Button(
                root,
                text="Назад",
                bg="#946A66",
                fg="white",
                activebackground="#946A66",
                activeforeground="white",
                relief="flat",
                command=lambda: create_user_home(user_fio, offset - 3),
            ).place(x=20, y=700)

            # Кнопка выхода с восстановлением оригинального стиля

        Button(
            root,
            text="Выход",
            command=logout,
            bg="#8B7355",
            fg="white",
            activebackground="#8B7355",
            activeforeground="white",
        ).place(x=width - 120, y=20)


    def logout():
        # Восстанавливаем оригинальные цвета
        global bg_color, fg_color
        bg_color = "#B18653"
        fg_color = "#C3A580"
        canvas.config(bg=bg_color)
        create_login_window()


    def create_personal_account(user_fio):
        """Создает страницу личного кабинета"""
        clear_window()

        # Устанавливаем цвет фона
        canvas.config(bg="#D5C5AE")

        # Создаем верхнюю полосу
        header_height = 60
        canvas.create_rectangle(0, 0, width, header_height, fill="#54101D", stipple="gray50", outline="")

        try:
            conn = psycopg2.connect(**DB_CONFIG)
            cursor = conn.cursor()

            # Получаем данные о читателе
            cursor.execute('SELECT "id", "Номер телефона", "Адрес" FROM "Читатели" WHERE "ФИО" = %s', (user_fio,))
            reader_id, phone, address = cursor.fetchone()

            # Получаем количество взятых книг
            cursor.execute(
                'SELECT COUNT(*) FROM "Состав выдачи" sv '
                'JOIN "Выдача" v ON sv."Выдача" = v."id" '
                'WHERE v."Читатель" = %s',
                (reader_id,),
            )
            total_books = cursor.fetchone()[0]

            # Получаем общее количество всех взятых книг (сумму количеств)
            cursor.execute(
                'SELECT COALESCE(SUM(sv."Количество"), 0) FROM "Состав выдачи" sv '
                'JOIN "Выдача" v ON sv."Выдача" = v."id" '
                'WHERE v."Читатель" = %s',
                (reader_id,),
            )
            total_books_sum = cursor.fetchone()[0]

            # Отображаем информацию
            info_y = header_height + 50
            info_width = 600

            # Создаем полупрозрачный белый блок для информации
            canvas.create_rectangle(
                (width - info_width) // 2,
                info_y,
                (width + info_width) // 2,
                info_y + 250,
                fill="white",
                stipple="gray75",
                outline="",
            )

            # Выводим информацию
            Label(root, text="Личный кабинет", font=("Arial", 16), bg="white", fg=text_color).place(
                x=width // 2, y=info_y + 20, anchor="center"
            )

            Label(root, text=f"ФИО: {user_fio}", bg="white", fg=text_color).place(
                x=width // 2 - info_width // 2 + 20, y=info_y + 60
            )

            Label(root, text=f"Номер телефона: {phone}", bg="white", fg=text_color).place(
                x=width // 2 - info_width // 2 + 20, y=info_y + 90
            )

            Label(root, text=f"Адрес: {address}", bg="white", fg=text_color).place(
                x=width // 2 - info_width // 2 + 20, y=info_y + 120
            )

            Label(root, text=f"Всего взято книг: {total_books}", bg="white", fg=text_color).place(
                x=width // 2 - info_width // 2 + 20, y=info_y + 150
            )

            Label(root, text=f"Общее количество экземпляров: {total_books_sum}", bg="white", fg=text_color).place(
                x=width // 2 - info_width // 2 + 20, y=info_y + 180
            )

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка получения данных: {str(e)}")
        finally:
            if "conn" in locals():
                conn.close()

        # Кнопка "Главная"
        Button(root, text="Главная", command=lambda: create_user_home(user_fio), bg="#8B7355", fg="white").place(x=20, y=20)

        # Кнопка выхода
        Button(
            root,
            text="Выход",
            command=logout,
            bg="#8B7355",
            fg="white",
            activebackground="#6B5A4D",
            activeforeground="white",
        ).place(x=width - 120, y=20)


    def create_staff_home(user_fio):
        """Создает главную страницу для сотрудника"""
        clear_window()

        # Устанавливаем цвет фона
        canvas.config(bg="#D5C5AE")

        # Создаем верхнюю полосу
        header_height = 60
        canvas.create_rectangle(0, 0, width, header_height, fill="#54101D", stipple="gray50", outline="")

        # Добавляем ФИО сотрудника по центру полосы
        Label(root, text=user_fio, font=("Arial", 16), bg="#54101D", fg="white").place(
            x=width // 2, y=header_height // 2, anchor="center"
        )

        # Создаем круг с инициалами (кнопка для личного кабинета)
        initials = "".join([name[0] for name in user_fio.split()[:2]])
        circle_radius = 25
        circle_x = width - 40
        circle_y = header_height // 2

        initials_btn = Button(
            root,
            text=initials,
            font=("Arial", 12),
            bg="white",
            fg="#54101D",
            bd=0,
            command=lambda: create_staff_account(user_fio),
        )
        initials_btn.place(
            x=circle_x - circle_radius, y=circle_y - circle_radius, width=circle_radius * 2, height=circle_radius * 2
        )

        # Создаем кнопки функционала сотрудника
        buttons = [
            {"text": "Добавить книгу", "command": lambda: create_add_book_window(user_fio)},
            {"text": "Редактировать книгу", "command": lambda: create_edit_book_window(user_fio)},
            {"text": "Удалить книгу", "command": lambda: create_delete_book_window(user_fio)},
            {"text": "Оформить выдачу", "command": lambda: create_issue_window(user_fio)},
            {"text": "Оформить возврат", "command": lambda: create_return_window(user_fio)},
            {"text": "Посмотреть список книг", "command": lambda: create_book_list_window(user_fio)},
            {"text": "Читатели с книгами", "command": lambda: create_readers_with_books_window(user_fio)},
            {"text": "Выгрузить данные", "command": lambda: create_export_window(user_fio)},
        ]

        # Расположение кнопок в сетке 2x4
        button_width = 200
        button_height = 50
        start_x = (width - 2 * (button_width + 20)) // 2
        start_y = header_height + 50

        for i, btn in enumerate(buttons):
            row = i // 2
            col = i % 2
            x = start_x + col * (button_width + 20)
            y = start_y + row * (button_height + 20)

            Button(
                root,
                text=btn["text"],
                command=btn["command"],
                bg="#8B7355",
                fg="white",
                width=button_width // 10,
                height=button_height // 20,
                activebackground="#6B5A4D",
                activeforeground="white",
            ).place(x=x, y=y)

        # Кнопка выхода
        Button(
            root,
            text="Выход",
            command=logout,
            bg="#8B7355",
            fg="white",
            activebackground="#6B5A4D",
            activeforeground="white",
        ).place(x=width - 120, y=20)


    def create_add_book_window(user_fio):
        """Создает окно для добавления новой книги"""
        clear_window()

        # Устанавливаем цвет фона
        canvas.config(bg="#D5C5AE")

        # Создаем верхнюю полосу
        header_height = 60
        canvas.create_rectangle(0, 0, width, header_height, fill="#54101D", stipple="gray50", outline="")

        # Добавляем ФИО сотрудника по центру полосы
        Label(root, text=user_fio, font=("Arial", 16), bg="#54101D", fg="white").place(
            x=width // 2, y=header_height // 2, anchor="center"
        )

        # Создаем круг с инициалами (кнопка для личного кабинета)
        initials = "".join([name[0] for name in user_fio.split()[:2]])
        circle_radius = 25
        circle_x = width - 40
        circle_y = header_height // 2

        initials_btn = Button(
            root,
            text=initials,
            font=("Arial", 12),
            bg="white",
            fg="#54101D",
            bd=0,
            command=lambda: create_staff_account(user_fio),
        )
        initials_btn.place(
            x=circle_x - circle_radius, y=circle_y - circle_radius, width=circle_radius * 2, height=circle_radius * 2
        )

        # Основной прямоугольник для формы
        form_width = 800
        form_height = 400
        form_x = (width - form_width) // 2
        form_y = header_height + 50

        round_rectangle(
            form_x, form_y, form_x + form_width, form_y + form_height, radius=20, fill="white", stipple="gray75"
        )

        # Левая часть формы - поля для ввода книги
        left_x = form_x + 50
        Label(root, text="Добавить книгу", font=("Arial", 14), bg="white", fg=text_color).place(x=left_x, y=form_y + 30)

        # Поля для ввода
        fields = [
            ("Название", form_y + 80),
            ("Автор", form_y + 130),
            ("Ценность", form_y + 180),
            ("Количество", form_y + 230),
        ]

        entries = []
        for text, y_pos in fields:
            Label(root, text=text + ":", bg="white", fg=text_color).place(x=left_x, y=y_pos)
            entry = Entry(root, width=30)
            entry.place(x=left_x + 100, y=y_pos)
            entries.append(entry)

        # Правая часть формы - информация о поставке
        right_x = form_x + form_width // 2 + 50

        # Получаем ID сотрудника из БД
        try:
            conn = psycopg2.connect(**DB_CONFIG)
            cursor = conn.cursor()
            cursor.execute('SELECT "id" FROM "Сотрудник" WHERE "ФИО" = %s', (user_fio,))
            staff_id = cursor.fetchone()[0]
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка получения ID сотрудника: {str(e)}")
            staff_id = "000000"
        finally:
            if "conn" in locals():
                conn.close()

        # Отображаем информацию о поставке
        Label(root, text=f"Id сотрудника: {staff_id}", bg="white", fg=text_color).place(x=right_x, y=form_y + 80)

        # Переменная для хранения даты
        current_date = datetime.date.today()
        date_var = StringVar()
        date_var.set(current_date.strftime("%d.%m.%Y"))

        Label(root, text="Дата:", bg="white", fg=text_color).place(x=right_x, y=form_y + 110)
        date_label = Label(root, textvariable=date_var, bg="white", fg=text_color)
        date_label.place(x=right_x + 50, y=form_y + 110)

        def change_date():
            """Создает окно для изменения даты"""
            date_window = Toplevel(root)
            date_window.title("Изменение даты")
            date_window.geometry("300x150")
            date_window.resizable(False, False)

            Label(date_window, text="Введите дату:", bg="white", fg=text_color).place(x=50, y=40)

            date_entry = Entry(date_window, width=15)
            date_entry.place(x=150, y=40)
            date_entry.insert(0, date_var.get())

            Label(date_window, text="В формате день.месяц.год", bg="white", fg="#777777").place(x=50, y=70)

            def apply_date():
                """Применяет новую дату"""
                try:
                    new_date = datetime.datetime.strptime(date_entry.get(), "%d.%m.%Y").date()
                    date_var.set(new_date.strftime("%d.%m.%Y"))
                    date_window.destroy()
                except ValueError:
                    messagebox.showerror("Ошибка", "Некорректный формат даты. Используйте ДД.ММ.ГГГГ")

            Button(date_window, text="Применить", command=apply_date, bg="#8B7355", fg="white").place(x=100, y=100)

        # Кнопка "Изменить дату"
        Button(
            root,
            text="Изменить дату",
            command=change_date,
            bg="#8B7355",
            fg="white",
            activebackground="#6B5A4D",
            activeforeground="white",
        ).place(x=right_x, y=form_y + 140)

        def add_book_to_db():
            """Добавляет книгу в базу данных"""
            title = entries[0].get()
            author = entries[1].get()
            value = entries[2].get()
            quantity = entries[3].get()

            if not all([title, author, value, quantity]):
                messagebox.showerror("Ошибка", "Все поля должны быть заполнены")
                return

            try:
                value = float(value)
                quantity = int(quantity)

                if quantity <= 0:
                    messagebox.showerror("Ошибка", "Количество должно быть положительным числом")
                    return

                supply_date = datetime.datetime.strptime(date_var.get(), "%d.%m.%Y").date()

                conn = psycopg2.connect(**DB_CONFIG)
                cursor = conn.cursor()

                # Добавляем книгу
                cursor.execute(
                    """
                    INSERT INTO "Книги" ("Название", "Автор", "Ценность", "Количество")
                    VALUES (%s, %s, %s, %s) RETURNING "id"
                """,
                    (title, author, value, quantity),
                )
                book_id = cursor.fetchone()[0]

                # Добавляем поставку
                cursor.execute(
                    """
                    INSERT INTO "Поставка" ("Сотрудник", "Дата")
                    VALUES (%s, %s) RETURNING "id"
                """,
                    (staff_id, supply_date),
                )
                supply_id = cursor.fetchone()[0]

                # Добавляем состав поставки
                cursor.execute(
                    """
                    INSERT INTO "Состав поставки" ("Поставка", "Книга", "Количество")
                    VALUES (%s, %s, %s)
                """,
                    (supply_id, book_id, quantity),
                )

                conn.commit()
                messagebox.showinfo("Успех", "Книга успешно добавлена!")
                create_staff_home(user_fio)

            except ValueError:
                messagebox.showerror("Ошибка", "Ценность и количество должны быть числами")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка при добавлении книги: {str(e)}")
                if "conn" in locals():
                    conn.rollback()
            finally:
                if "conn" in locals():
                    conn.close()

        # Кнопка "Добавить книгу"
        Button(
            root,
            text="Добавить книгу",
            command=add_book_to_db,
            bg="#8B7355",
            fg="white",
            activebackground="#6B5A4D",
            activeforeground="white",
        ).place(x=form_x + form_width // 2 - 50, y=form_y + form_height - 50)

        # Кнопка "Назад"
        Button(
            root,
            text="Назад",
            command=lambda: create_staff_home(user_fio),
            bg="#8B7355",
            fg="white",
            activebackground="#6B5A4D",
            activeforeground="white",
        ).place(x=20, y=20)

        # Кнопка выхода
        Button(
            root,
            text="Выход",
            command=logout,
            bg="#8B7355",
            fg="white",
            activebackground="#6B5A4D",
            activeforeground="white",
        ).place(x=width - 120, y=20)


    def create_staff_account(user_fio):
        """Создает страницу личного кабинета сотрудника"""
        clear_window()

        # Устанавливаем цвет фона
        canvas.config(bg="#D5C5AE")

        # Создаем верхнюю полосу
        header_height = 60
        canvas.create_rectangle(0, 0, width, header_height, fill="#54101D", stipple="gray50", outline="")

        try:
            conn = psycopg2.connect(**DB_CONFIG)
            cursor = conn.cursor()

            # Получаем данные о сотруднике
            cursor.execute('SELECT "id", "Номер телефона" FROM "Сотрудник" WHERE "ФИО" = %s', (user_fio,))
            staff_id, phone = cursor.fetchone()

            # Отображаем информацию
            info_y = header_height + 50
            info_width = 600

            # Создаем полупрозрачный белый блок для информации
            canvas.create_rectangle(
                (width - info_width) // 2,
                info_y,
                (width + info_width) // 2,
                info_y + 200,
                fill="white",
                stipple="gray75",
                outline="",
            )

            # Выводим информацию
            Label(root, text="Личный кабинет сотрудника", font=("Arial", 16), bg="white", fg=text_color).place(
                x=width // 2, y=info_y + 20, anchor="center"
            )

            Label(root, text=f"ФИО: {user_fio}", bg="white", fg=text_color).place(
                x=width // 2 - info_width // 2 + 20, y=info_y + 60
            )

            Label(root, text=f"Номер телефона: {phone}", bg="white", fg=text_color).place(
                x=width // 2 - info_width // 2 + 20, y=info_y + 90
            )

            Label(root, text=f"ID сотрудника: {staff_id}", bg="white", fg=text_color).place(
                x=width // 2 - info_width // 2 + 20, y=info_y + 120
            )

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка получения данных: {str(e)}")
        finally:
            if "conn" in locals():
                conn.close()

        # Кнопка "Главная"
        Button(root, text="Главная", command=lambda: create_staff_home(user_fio), bg="#8B7355", fg="white").place(
            x=20, y=20
        )

        # Кнопка выхода
        Button(root, text="Выход", command=logout, bg="#8B7355", fg="white").place(x=width - 120, y=20)


    def create_edit_book_window(user_fio, page=0, sort_by="name", sort_order="asc"):
        """Создает окно для редактирования книг"""
        clear_window()

        # Устанавливаем цвет фона
        canvas.config(bg="#D5C5AE")

        # Создаем верхнюю полосу
        header_height = 60
        canvas.create_rectangle(0, 0, width, header_height, fill="#54101D", stipple="gray50", outline="")

        # Добавляем ФИО сотрудника по центру полосы
        Label(root, text=user_fio, font=("Arial", 16), bg="#54101D", fg="white").place(
            x=width // 2, y=header_height // 2, anchor="center"
        )

        # Создаем круг с инициалами
        initials = "".join([name[0] for name in user_fio.split()[:2]])
        circle_radius = 25
        circle_x = width - 40
        circle_y = header_height // 2

        initials_btn = Button(
            root,
            text=initials,
            font=("Arial", 12),
            bg="white",
            fg="#54101D",
            bd=0,
            command=lambda: create_staff_account(user_fio),
        )
        initials_btn.place(
            x=circle_x - circle_radius, y=circle_y - circle_radius, width=circle_radius * 2, height=circle_radius * 2
        )

        # Заголовок окна
        Label(root, text="Редактировать книгу", font=("Arial", 14), bg="#D5C5AE", fg="#54101D").place(
            x=width // 2, y=header_height + 30, anchor="center"
        )

        # Получаем список книг из БД
        try:
            conn = psycopg2.connect(**DB_CONFIG)
            cursor = conn.cursor()

            # Определяем порядок сортировки
            order_by = {"name": '"Название"', "value": '"Ценность"', "quantity": '"Количество"'}.get(sort_by, '"Название"')

            order_dir = "ASC" if sort_order == "asc" else "DESC"

            cursor.execute(
                f"""
                SELECT 
                    k."id", 
                    k."Название", 
                    k."Автор", 
                    k."Ценность",
                    COALESCE(
                        (
                            SELECT SUM(sp."Количество")
                            FROM "Состав поставки" sp
                            WHERE sp."Книга" = k."id"
                        ) - 
                        COALESCE(
                            (
                                SELECT SUM(sv."Количество")
                                FROM "Состав выдачи" sv
                                JOIN "Выдача" v ON sv."Выдача" = v."id"
                                WHERE sv."Книга" = k."id" AND v."Создана" = FALSE
                            ), 0
                        ), 0
                    ) as "Количество"
                FROM "Книги" k
                ORDER BY {order_by} {order_dir}
            """
            )
            all_books = cursor.fetchall()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка получения списка книг: {str(e)}")
            all_books = []
        finally:
            if "conn" in locals():
                conn.close()

        # Разбиваем на страницы (по 4 книги на страницу)
        books_per_page = 4
        total_pages = max(1, (len(all_books) + books_per_page - 1) // books_per_page)
        current_page = page % total_pages
        start_idx = current_page * books_per_page
        end_idx = min(start_idx + books_per_page, len(all_books))
        books_to_show = all_books[start_idx:end_idx]

        # Основной контейнер для списка книг
        list_frame = Frame(root, bg="#D5C5AE")
        list_frame.place(x=width // 2 - 300, y=header_height + 100, width=600, height=500)

        # Варианты сортировки
        sort_frame = Frame(root, bg="#D5C5AE")
        sort_frame.place(x=width // 2 - 300, y=header_height + 60, width=600, height=30)

        Label(sort_frame, text="Сортировать:", bg="#D5C5AE", fg="#54101D").pack(side=LEFT)

        # Кнопки сортировки
        sort_options = [("По алфавиту", "name"), ("По количеству", "quantity"), ("По ценности", "value")]

        for text, sort_key in sort_options:
            # Создаем отдельную функцию для каждой кнопки сортировки
            def make_sort_func(key):
                def sort_func():
                    # Определяем новый порядок сортировки
                    new_order = "desc" if sort_by == key and sort_order == "asc" else "asc"
                    create_edit_book_window(user_fio, current_page, key, new_order)

                return sort_func

            btn = Button(
                sort_frame,
                text=text,
                command=make_sort_func(sort_key),
                bg="#8B7355",
                fg="white",
                activebackground="#6B5A4D",
                activeforeground="white",
            )
            btn.pack(side=LEFT, padx=5)

        # Отображаем книги
        for i, (book_id, title, author, value, quantity) in enumerate(books_to_show):
            book_frame = Frame(list_frame, bg="white")
            book_frame.pack(fill=X, pady=5)

            # Создаем фон прямоугольника
            book_canvas = Canvas(book_frame, width=1200, height=80, bg="white", highlightthickness=0)
            book_canvas.pack()

            # Рисуем прямоугольник
            book_canvas.create_rectangle(0, 0, 1200, 80, fill="white", stipple="gray50", outline="")

            # Добавляем текст
            book_canvas.create_text(10, 15, text=author, anchor="nw", fill="#54101D", font=("Arial", 10))
            book_canvas.create_text(1100, 15, text=title, anchor="ne", fill="#54101D", font=("Arial", 10, "bold"))
            book_canvas.create_text(
                215, 45, text=f"Ценность: {value} | Количество: {quantity}", anchor="nw", fill="#54101D", font=("Arial", 10)
            )

            # Кнопка редактирования
            Button(
                book_frame,
                text="Редактировать",
                bg="#8B7355",
                fg="white",
                command=lambda bid=book_id: edit_book(bid, user_fio),
            ).pack(pady=5)

            # Кнопки навигации
        nav_frame = Frame(root, bg="#D5C5AE")
        nav_frame.place(x=width // 2 - 75, y=header_height + 640, width=200, height=40)

        Button(
            nav_frame,
            text="Назад",
            command=lambda: create_edit_book_window(user_fio, current_page - 1, sort_by, sort_order),
            bg="#8B7355",
            fg="white",
            activebackground="#6B5A4D",
            activeforeground="white",
        ).pack(side=LEFT, padx=10)

        Button(
            nav_frame,
            text="Далее",
            command=lambda: create_edit_book_window(user_fio, current_page + 1, sort_by, sort_order),
            bg="#8B7355",
            fg="white",
            activebackground="#6B5A4D",
            activeforeground="white",
        ).pack(side=LEFT, padx=10)

        # Кнопка "Назад" в главное меню
        Button(
            root,
            text="Назад",
            command=lambda: create_staff_home(user_fio),
            bg="#8B7355",
            fg="white",
            activebackground="#6B5A4D",
            activeforeground="white",
        ).place(x=20, y=20)

        # Кнопка выхода
        Button(
            root,
            text="Выход",
            command=logout,
            bg="#8B7355",
            fg="white",
            activebackground="#6B5A4D",
            activeforeground="white",
        ).place(x=width - 120, y=20)


    def edit_book(book_id, user_fio):
        """Окно редактирования книги"""
        edit_win = Toplevel(root)
        edit_win.title("Редактировать книгу")
        edit_win.geometry("600x400")
        edit_win.config(bg="#D5C5AE")

        # Получаем текущие данные книги из БД
        conn = psycopg2.connect(**DB_CONFIG)
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT "Название", "Автор", "Ценность"
            FROM "Книги"
            WHERE "id" = %s
        """,
            (book_id,),
        )
        current_data = cursor.fetchone()
        conn.close()

        # Поля для редактирования
        fields = [
            ("Название", current_data[0]),
            ("Автор", current_data[1]),
            ("Ценность", current_data[2]),
            ("Количество", get_book_quantity(book_id)),
        ]

        entries = []
        y_pos = 100
        for label_text, initial_value in fields:
            Label(edit_win, text=label_text, bg="#D5C5AE", fg=text_color).place(x=150, y=y_pos)
            entry = Entry(edit_win, width=40)
            entry.insert(0, str(initial_value))  # Явное преобразование в строку
            entry.place(x=250, y=y_pos)
            entries.append(entry)
            y_pos += 40

        def save_changes():
            """Сохраняет изменения в базе данных"""
            try:
                # Получаем значения из полей ввода
                new_title = entries[0].get()
                new_author = entries[1].get()
                new_value = float(entries[2].get())
                new_quantity = int(entries[3].get())

                # Проверка на отрицательные значения
                if new_quantity < 0:
                    messagebox.showerror("Ошибка", "Количество не может быть отрицательным")
                    return

                # Соединяемся с БД и обновляем данные
                conn = psycopg2.connect(**DB_CONFIG)
                cursor = conn.cursor()

                # Получаем текущее количество выданных книг
                cursor.execute(
                    """
                    SELECT COALESCE(SUM(sv."Количество"), 0)
                    FROM "Состав выдачи" sv
                    JOIN "Выдача" v ON sv."Выдача" = v."id"
                    WHERE sv."Книга" = %s AND v."Создана" = FALSE
                """,
                    (book_id,),
                )
                issued_quantity = cursor.fetchone()[0]

                # Проверяем, что новое количество не меньше выданных книг
                if new_quantity < issued_quantity:
                    messagebox.showerror(
                        "Ошибка",
                        f"Новое количество ({new_quantity}) не может быть меньше количества выданных книг ({issued_quantity})",
                    )
                    return

                # Обновляем основные данные книги
                cursor.execute(
                    """
                    UPDATE "Книги"
                    SET "Название" = %s, "Автор" = %s, "Ценность" = %s, "Количество" = %s
                    WHERE "id" = %s
                """,
                    (new_title, new_author, new_value, new_quantity, book_id),
                )

                # Обновляем количество в составе поставки
                cursor.execute(
                    """
                    UPDATE "Состав поставки"
                    SET "Количество" = %s
                    WHERE "Книга" = %s
                """,
                    (new_quantity, book_id),
                )

                conn.commit()
                conn.close()

                # Закрываем окно редактирования и обновляем список
                edit_win.destroy()
                create_edit_book_window(user_fio)
                messagebox.showinfo("Успех", "Изменения успешно сохранены!")

            except ValueError:
                messagebox.showerror("Ошибка", "Ценность и количество должны быть числами")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка сохранения: {str(e)}")
                if "conn" in locals():
                    conn.rollback()

        # Кнопка сохранения
        Button(edit_win, text="Сохранить", command=save_changes, bg="#8B7355", fg="white").place(x=400, y=300)


    def get_book_quantity(book_id):
        """Получает текущее количество книги"""
        conn = psycopg2.connect(**DB_CONFIG)
        cursor = conn.cursor()

        # Суммируем поставки и вычитаем выданные книги (только не возвращенные)
        cursor.execute(
            """
            SELECT 
                COALESCE(SUM(sp."Количество"), 0) - 
                COALESCE((
                    SELECT SUM(sv."Количество")
                    FROM "Состав выдачи" sv
                    JOIN "Выдача" v ON sv."Выдача" = v."id"
                    WHERE sv."Книга" = %s AND v."Создана" = FALSE
                ), 0)
            FROM 
                "Состав поставки" sp
            WHERE 
                sp."Книга" = %s
        """,
            (book_id, book_id),
        )

        quantity = cursor.fetchone()[0]
        conn.close()
        return quantity


    def create_delete_book_window(user_fio, page=0, sort_by="name", sort_order="asc"):
        """Создает окно для удаления книг"""
        clear_window()
        canvas.config(bg="#D5C5AE")
        header_height = 60
        canvas.create_rectangle(0, 0, width, header_height, fill="#54101D", stipple="gray50", outline="")
        Label(root, text=user_fio, font=("Arial", 16), bg="#54101D", fg="white").place(
            x=width // 2, y=header_height // 2, anchor="center"
        )
        initials = "".join([name[0] for name in user_fio.split()[:2]])
        circle_radius = 25
        circle_x = width - 40
        circle_y = header_height // 2
        initials_btn = Button(
            root,
            text=initials,
            font=("Arial", 12),
            bg="white",
            fg="#54101D",
            bd=0,
            command=lambda: create_staff_account(user_fio),
        )
        initials_btn.place(
            x=circle_x - circle_radius, y=circle_y - circle_radius, width=circle_radius * 2, height=circle_radius * 2
        )
        Label(root, text="Удалить книгу", font=("Arial", 16), bg="#D5C5AE", fg="#54101D").place(
            x=width // 2, y=header_height + 30, anchor="center"
        )
        # Получаем список книг
        try:
            conn = psycopg2.connect(**DB_CONFIG)
            cursor = conn.cursor()
            order_by = {"name": '"Название"', "value": '"Ценность"', "quantity": '"Количество"'}.get(sort_by, '"Название"')
            order_dir = "ASC" if sort_order == "asc" else "DESC"
            cursor.execute(
                f"""
                SELECT 
                    k."id", 
                    k."Название", 
                    k."Автор", 
                    k."Ценность",
                    COALESCE(
                        (
                            SELECT SUM(sp."Количество")
                            FROM "Состав поставки" sp
                            WHERE sp."Книга" = k."id"
                        ) - 
                        COALESCE(
                            (
                                SELECT SUM(sv."Количество")
                                FROM "Состав выдачи" sv
                                JOIN "Выдача" v ON sv."Выдача" = v."id"
                                WHERE sv."Книга" = k."id" AND v."Создана" = FALSE
                            ), 0
                        ), 0
                    ) as "Количество"
                FROM "Книги" k
                ORDER BY {order_by} {order_dir}
            """
            )
            all_books = cursor.fetchall()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка получения списка книг: {str(e)}")
            all_books = []
        finally:
            if "conn" in locals():
                conn.close()
        books_per_page = 4
        total_pages = max(1, (len(all_books) + books_per_page - 1) // books_per_page)
        current_page = page % total_pages
        start_idx = current_page * books_per_page
        end_idx = min(start_idx + books_per_page, len(all_books))
        books_to_show = all_books[start_idx:end_idx]
        list_frame = Frame(root, bg="#D5C5AE")
        list_frame.place(x=width // 2 - 300, y=header_height + 100, width=600, height=500)
        sort_frame = Frame(root, bg="#D5C5AE")
        sort_frame.place(x=width // 2 - 300, y=header_height + 60, width=600, height=30)
        Label(sort_frame, text="Сортировать:", bg="#D5C5AE", fg="#54101D").pack(side=LEFT)
        sort_options = [("По алфавиту", "name"), ("По количеству", "quantity"), ("По ценности", "value")]
        for text, sort_key in sort_options:

            def make_sort_func(key):
                def sort_func():
                    new_order = "desc" if sort_by == key and sort_order == "asc" else "asc"
                    create_delete_book_window(user_fio, current_page, key, new_order)

                return sort_func

            btn = Button(
                sort_frame,
                text=text,
                command=make_sort_func(sort_key),
                bg="#8B7355",
                fg="white",
                activebackground="#6B5A4D",
                activeforeground="white",
            )
            btn.pack(side=LEFT, padx=5)
        for i, (book_id, title, author, value, quantity) in enumerate(books_to_show):
            book_frame = Frame(list_frame, bg="white")
            book_frame.pack(fill=X, pady=5)
            book_canvas = Canvas(book_frame, width=1200, height=80, bg="white", highlightthickness=0)
            book_canvas.pack()
            book_canvas.create_rectangle(0, 0, 1200, 80, fill="white", stipple="gray50", outline="")
            book_canvas.create_text(10, 15, text=author, anchor="nw", fill="#54101D", font=("Arial", 10))
            book_canvas.create_text(1100, 15, text=title, anchor="ne", fill="#54101D", font=("Arial", 10, "bold"))
            book_canvas.create_text(
                215, 45, text=f"Ценность: {value} | Количество: {quantity}", anchor="nw", fill="#54101D", font=("Arial", 10)
            )
            Button(
                book_frame,
                text="Удалить",
                bg="#8B7355",
                fg="white",
                command=lambda bid=book_id: confirm_delete_book(bid, user_fio, page, sort_by, sort_order),
            ).pack(pady=5)
        nav_frame = Frame(root, bg="#D5C5AE")
        nav_frame.place(x=width // 2 - 75, y=header_height + 640, width=200, height=40)
        Button(
            nav_frame,
            text="Назад",
            command=lambda: create_delete_book_window(user_fio, current_page - 1, sort_by, sort_order),
            bg="#8B7355",
            fg="white",
            activebackground="#6B5A4D",
            activeforeground="white",
        ).pack(side=LEFT, padx=10)
        Button(
            nav_frame,
            text="Далее",
            command=lambda: create_delete_book_window(user_fio, current_page + 1, sort_by, sort_order),
            bg="#8B7355",
            fg="white",
            activebackground="#6B5A4D",
            activeforeground="white",
        ).pack(side=LEFT, padx=10)
        Button(
            root,
            text="Назад",
            command=lambda: create_staff_home(user_fio),
            bg="#8B7355",
            fg="white",
            activebackground="#6B5A4D",
            activeforeground="white",
        ).place(x=20, y=20)
        Button(
            root,
            text="Выход",
            command=logout,
            bg="#8B7355",
            fg="white",
            activebackground="#6B5A4D",
            activeforeground="white",
        ).place(x=width - 120, y=20)


    def confirm_delete_book(book_id, user_fio, page, sort_by, sort_order):
        """Всплывающее окно подтверждения удаления книги"""
        confirm_win = Toplevel(root)
        confirm_win.title("Подтверждение удаления")
        confirm_win.geometry("350x150")
        confirm_win.config(bg="#E5D3C6")
        Label(
            confirm_win, text="Вы уверены, что хотите удалить эту книгу?", font=("Arial", 12), bg="#E5D3C6", fg="#54101D"
        ).place(x=30, y=20, width=290, height=50)
        # Кнопки Да и Нет размещаем через place, без рамок
        btn_yes = Button(
            confirm_win,
            text="Да",
            command=lambda: do_delete(),
            bg="#8B7355",
            fg="white",
            width=10,
            bd=0,
            highlightthickness=0,
        )
        btn_no = Button(
            confirm_win,
            text="Нет",
            command=confirm_win.destroy,
            bg="#8B7355",
            fg="white",
            width=10,
            bd=0,
            highlightthickness=0,
        )
        btn_yes.place(x=70, y=80, width=80, height=35)
        btn_no.place(x=200, y=80, width=80, height=35)

        def do_delete():
            try:
                conn = psycopg2.connect(**DB_CONFIG)
                cursor = conn.cursor()
                cursor.execute('DELETE FROM "Состав выдачи" WHERE "Книга" = %s', (book_id,))
                cursor.execute('DELETE FROM "Состав поставки" WHERE "Книга" = %s', (book_id,))
                cursor.execute('DELETE FROM "Книги" WHERE "id" = %s', (book_id,))
                conn.commit()
                conn.close()
                confirm_win.destroy()
                messagebox.showinfo("Успех", "Книга успешно удалена!")
                create_delete_book_window(user_fio, page, sort_by, sort_order)
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка при удалении: {str(e)}")
                if "conn" in locals():
                    conn.rollback()


    def create_issue_window(user_fio):
        """Окно оформления выдачи книги (шаг 1)"""
        clear_window()
        canvas.config(bg="#D5C5AE")
        header_height = 60
        canvas.create_rectangle(0, 0, width, header_height, fill="#54101D", stipple="gray50", outline="")
        Label(root, text=user_fio, font=("Arial", 16), bg="#54101D", fg="white").place(
            x=width // 2, y=header_height // 2, anchor="center"
        )
        initials = "".join([name[0] for name in user_fio.split()[:2]])
        circle_radius = 25
        circle_x = width - 40
        circle_y = header_height // 2
        initials_btn = Button(
            root,
            text=initials,
            font=("Arial", 12),
            bg="white",
            fg="#54101D",
            bd=0,
            command=lambda: create_staff_account(user_fio),
        )
        initials_btn.place(
            x=circle_x - circle_radius, y=circle_y - circle_radius, width=circle_radius * 2, height=circle_radius * 2
        )
        Label(root, text="Выдача", font=("Arial", 16), bg="#D5C5AE", fg="#54101D").place(
            x=width // 2, y=header_height + 30, anchor="center"
        )
        # Фон формы
        form_width = 800
        form_height = 400
        form_x = (width - form_width) // 2
        form_y = header_height + 50
        round_rectangle(
            form_x, form_y, form_x + form_width, form_y + form_height, radius=50, fill="#D2B28F", stipple="gray50"
        )
        # Поле ФИО
        fio_entry = Entry(root, font=("Arial", 14), bg="#D2B28F", fg="#2A1E17", bd=2, relief="flat")
        fio_entry.place(x=form_x + 100, y=form_y + 70, width=600, height=40)
        fio_entry.insert(0, "ФИО читателя")
        # Даты
        today = datetime.date.today()
        one_month = today + datetime.timedelta(days=30)
        date_issue_var = StringVar()
        date_return_var = StringVar()
        date_issue_var.set(today.strftime("%d.%m.%Y"))
        date_return_var.set(one_month.strftime("%d.%m.%Y"))
        # Дата выдачи
        date_issue_entry = Entry(
            root, textvariable=date_issue_var, font=("Arial", 14), bg="#D2B28F", fg="#2A1E17", bd=2, relief="flat"
        )
        date_issue_entry.place(x=form_x + 100, y=form_y + 140, width=250, height=40)
        # Дата возврата
        date_return_entry = Entry(
            root, textvariable=date_return_var, font=("Arial", 14), bg="#D2B28F", fg="#2A1E17", bd=2, relief="flat"
        )
        date_return_entry.place(x=form_x + 450, y=form_y + 140, width=250, height=40)
        # Кнопка Далее
        Button(
            root,
            text="Далее",
            font=("Arial", 14),
            bg="#E5D3C6",
            fg="#2A1E17",
            bd=2,
            relief="solid",
            command=lambda: create_issue_step2_window(
                user_fio, fio_entry.get(), date_issue_var.get(), date_return_var.get()
            ),
        ).place(x=form_x + 250, y=form_y + 230, width=200, height=60)
        # Кнопка Назад
        Button(
            root,
            text="Назад",
            command=lambda: create_staff_home(user_fio),
            bg="#8B7355",
            fg="white",
            activebackground="#6B5A4D",
            activeforeground="white",
        ).place(x=20, y=20)
        # Кнопка Выход
        Button(
            root,
            text="Выход",
            command=logout,
            bg="#8B7355",
            fg="white",
            activebackground="#6B5A4D",
            activeforeground="white",
        ).place(x=width - 120, y=20)


    def create_issue_step2_window(user_fio, fio, date_issue, date_return):
        """Окно выбора книги и количества (шаг 2)"""
        clear_window()
        canvas.config(bg="#D5C5AE")
        header_height = 60
        canvas.create_rectangle(0, 0, width, header_height, fill="#54101D", stipple="gray50", outline="")
        Label(root, text=user_fio, font=("Arial", 16), bg="#54101D", fg="white").place(
            x=width // 2, y=header_height // 2, anchor="center"
        )
        initials = "".join([name[0] for name in user_fio.split()[:2]])
        circle_radius = 25
        circle_x = width - 40
        circle_y = header_height // 2
        initials_btn = Button(
            root,
            text=initials,
            font=("Arial", 12),
            bg="white",
            fg="#54101D",
            bd=0,
            command=lambda: create_staff_account(user_fio),
        )
        initials_btn.place(
            x=circle_x - circle_radius, y=circle_y - circle_radius, width=circle_radius * 2, height=circle_radius * 2
        )
        Label(root, text="Выдача", font=("Arial", 16), bg="#D5C5AE", fg="#54101D").place(
            x=width // 2, y=header_height + 30, anchor="center"
        )
        # Фон формы
        form_width = 800
        form_height = 400
        form_x = (width - form_width) // 2
        form_y = header_height + 50
        round_rectangle(
            form_x, form_y, form_x + form_width, form_y + form_height, radius=50, fill="#D2B28F", stipple="gray50"
        )
        # Получаем список книг из БД
        try:
            conn = psycopg2.connect(**DB_CONFIG)
            cursor = conn.cursor()
            cursor.execute('SELECT "Название" FROM "Книги" ORDER BY "Название" ASC')
            books = [row[0] for row in cursor.fetchall()]
            conn.close()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка получения списка книг: {str(e)}")
            books = []
        # Поле Название книги с выпадающим списком
        book_var = StringVar()
        book_entry = Entry(root, textvariable=book_var, font=("Arial", 14), bg="#D2B28F", fg="#2A1E17", bd=2, relief="flat")
        book_entry.place(x=form_x + 100, y=form_y + 70, width=600, height=40)
        book_entry.insert(0, "Название книги")

        # Кнопка для показа списка
        def show_book_list():
            list_win = Toplevel(root)
            list_win.title("Выберите книгу")
            list_win.geometry("400x300")
            listbox = Listbox(list_win, font=("Arial", 12))
            for b in books:
                listbox.insert("end", b)
            listbox.pack(fill="both", expand=True)

            def select_book(event):
                selection = listbox.get(listbox.curselection())
                book_var.set(selection)
                list_win.destroy()

            listbox.bind("<Double-Button-1>", select_book)

        Button(root, text="▼", font=("Arial", 12), command=show_book_list, bg="#8B7355", fg="white", bd=0).place(
            x=form_x + 710, y=form_y + 70, width=40, height=40
        )
        # Поле Количество
        quantity_entry = Entry(root, font=("Arial", 14), bg="#D2B28F", fg="#2A1E17", bd=2, relief="flat")
        quantity_entry.place(x=form_x + 100, y=form_y + 140, width=600, height=40)
        quantity_entry.insert(0, "Количество")
        # Кнопка Готово
        Button(
            root,
            text="Готово",
            font=("Arial", 14),
            bg="#E5D3C6",
            fg="#2A1E17",
            bd=2,
            relief="solid",
            command=lambda: save_issue(user_fio, fio, date_issue, date_return, book_var.get(), quantity_entry.get()),
        ).place(x=form_x + 250, y=form_y + 230, width=200, height=60)
        # Кнопка Назад
        Button(
            root,
            text="Назад",
            command=lambda: create_issue_window(user_fio),
            bg="#8B7355",
            fg="white",
            activebackground="#6B5A4D",
            activeforeground="white",
        ).place(x=20, y=20)
        # Кнопка Выход
        Button(
            root,
            text="Выход",
            command=logout,
            bg="#8B7355",
            fg="white",
            activebackground="#6B5A4D",
            activeforeground="white",
        ).place(x=width - 120, y=20)


    def save_issue(user_fio, fio, date_issue, date_return, book_title, quantity):
        """Сохраняет выдачу в БД и возвращает на главную"""
        try:
            conn = psycopg2.connect(**DB_CONFIG)
            cursor = conn.cursor()
            # Получаем id читателя
            cursor.execute('SELECT "id" FROM "Читатели" WHERE "ФИО" = %s', (fio,))
            reader = cursor.fetchone()
            if not reader:
                messagebox.showerror("Ошибка", "Читатель не найден!")
                return
            reader_id = reader[0]
            # Получаем id сотрудника
            cursor.execute('SELECT "id" FROM "Сотрудник" WHERE "ФИО" = %s', (user_fio,))
            staff = cursor.fetchone()
            if not staff:
                messagebox.showerror("Ошибка", "Сотрудник не найден!")
                return
            staff_id = staff[0]
            # Получаем id книги
            cursor.execute('SELECT "id" FROM "Книги" WHERE "Название" = %s', (book_title,))
            book = cursor.fetchone()
            if not book:
                messagebox.showerror("Ошибка", "Книга не найдена!")
                return
            book_id = book[0]
            # Проверяем количество
            try:
                quantity = int(quantity)
                if quantity <= 0:
                    raise ValueError
            except ValueError:
                messagebox.showerror("Ошибка", "Количество должно быть положительным числом!")
                return
                # Проверяем, достаточно ли книг
                cursor.execute(
                    """
                SELECT COALESCE(
                    (SELECT SUM(sp."Количество") FROM "Состав поставки" sp WHERE sp."Книга" = %s) -
                    COALESCE((SELECT SUM(sv."Количество") FROM "Состав выдачи" sv JOIN "Выдача" v ON sv."Выдача" = v."id" WHERE sv."Книга" = %s AND v."Создана" = FALSE), 0), 0)
            """,
                    (book_id, book_id),
                )
            available = cursor.fetchone()[0]
            if quantity > available:
                messagebox.showerror("Ошибка", f"Недостаточно экземпляров книги! Доступно: {available}")
                return
            # Создаём выдачу
            cursor.execute(
                """
                INSERT INTO "Выдача" ("Сотрудник", "Читатель", "Дата выдачи", "Дата возврата", "Создана")
                VALUES (%s, %s, %s, %s, FALSE) RETURNING "id"
            """,
                (
                    staff_id,
                    reader_id,
                    datetime.datetime.strptime(date_issue, "%d.%m.%Y").date(),
                    datetime.datetime.strptime(date_return, "%d.%m.%Y").date(),
                ),
            )
            issue_id = cursor.fetchone()[0]
            # Добавляем в состав выдачи
            cursor.execute(
                """
                INSERT INTO "Состав выдачи" ("Выдача", "Книга", "Количество")
                VALUES (%s, %s, %s)
            """,
                (issue_id, book_id, quantity),
            )
            conn.commit()
            conn.close()
            messagebox.showinfo("Успех", "Выдача успешно оформлена!")
            create_staff_home(user_fio)
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при оформлении выдачи: {str(e)}")


    def create_return_window(user_fio):
        """Окно оформления возврата книги"""
        clear_window()
        canvas.config(bg="#D5C5AE")
        header_height = 60
        canvas.create_rectangle(0, 0, width, header_height, fill="#54101D", stipple="gray50", outline="")
        Label(root, text=user_fio, font=("Arial", 16), bg="#54101D", fg="white").place(
            x=width // 2, y=header_height // 2, anchor="center"
        )
        initials = "".join([name[0] for name in user_fio.split()[:2]])
        circle_radius = 25
        circle_x = width - 40
        circle_y = header_height // 2
        initials_btn = Button(
            root,
            text=initials,
            font=("Arial", 12),
            bg="white",
            fg="#54101D",
            bd=0,
            command=lambda: create_staff_account(user_fio),
        )
        initials_btn.place(
            x=circle_x - circle_radius, y=circle_y - circle_radius, width=circle_radius * 2, height=circle_radius * 2
        )
        Label(root, text="Возврат", font=("Arial", 16), bg="#D5C5AE", fg="#54101D").place(
            x=width // 2, y=header_height + 30, anchor="center"
        )
        # Фон формы
        form_width = 800
        form_height = 400
        form_x = (width - form_width) // 2
        form_y = header_height + 50
        round_rectangle(
            form_x, form_y, form_x + form_width, form_y + form_height, radius=50, fill="#D2B28F", stipple="gray50"
        )
        # Получаем список читателей
        try:
            conn = psycopg2.connect(**DB_CONFIG)
            cursor = conn.cursor()
            cursor.execute('SELECT "ФИО" FROM "Читатели" ORDER BY "ФИО" ASC')
            readers = [row[0] for row in cursor.fetchall()]
            conn.close()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка получения читателей: {str(e)}")
            readers = []
        reader_var = StringVar()
        reader_entry = Entry(
            root, textvariable=reader_var, font=("Arial", 14), bg="#D2B28F", fg="#2A1E17", bd=2, relief="flat"
        )
        reader_entry.place(x=form_x + 100, y=form_y + 70, width=600, height=40)
        reader_entry.insert(0, "ФИО читателя")

        def show_reader_list():
            list_win = Toplevel(root)
            list_win.title("Выберите читателя")
            list_win.geometry("400x300")
            listbox = Listbox(list_win, font=("Arial", 12))
            for r in readers:
                listbox.insert("end", r)
            listbox.pack(fill="both", expand=True)

            def select_reader(event):
                selection = listbox.get(listbox.curselection())
                reader_var.set(selection)
                list_win.destroy()
                update_issues_list()

            listbox.bind("<Double-Button-1>", select_reader)

        Button(root, text="▼", font=("Arial", 12), command=show_reader_list, bg="#8B7355", fg="white", bd=0).place(
            x=form_x + 710, y=form_y + 70, width=40, height=40
        )
        # Список выдач
        issue_var = StringVar()
        issue_entry = Entry(
            root, textvariable=issue_var, font=("Arial", 14), bg="#D2B28F", fg="#2A1E17", bd=2, relief="flat"
        )
        issue_entry.place(x=form_x + 100, y=form_y + 140, width=600, height=40)
        issue_entry.insert(0, "Выдача")

        def show_issues_list():
            fio = reader_var.get()
            if not fio or fio == "ФИО читателя":
                messagebox.showerror("Ошибка", "Сначала выберите читателя!")
                return
            try:
                conn = psycopg2.connect(**DB_CONFIG)
                cursor = conn.cursor()
                cursor.execute(
                    """
                    SELECT v."id", v."Дата выдачи", v."Дата возврата" FROM "Выдача" v
                    JOIN "Читатели" c ON v."Читатель" = c."id"
                    WHERE c."ФИО" = %s AND v."Создана" = FALSE
                    ORDER BY v."Дата выдачи" ASC
                """,
                    (fio,),
                )
                issues = cursor.fetchall()
                conn.close()
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка получения выдач: {str(e)}")
                return
            if not issues:
                messagebox.showinfo("Нет выдач", "У этого читателя нет не возвращённых выдач!")
                return
            list_win = Toplevel(root)
            list_win.title("Выберите выдачу")
            list_win.geometry("400x300")
            listbox = Listbox(list_win, font=("Arial", 12))
            for v_id, d1, d2 in issues:
                listbox.insert("end", f"ID {v_id} | {d1.strftime('%d.%m.%Y')} - {d2.strftime('%d.%m.%Y')}")
            listbox.pack(fill="both", expand=True)

            def select_issue(event):
                selection = listbox.get(listbox.curselection())
                issue_var.set(selection)
                list_win.destroy()

            listbox.bind("<Double-Button-1>", select_issue)

        def update_issues_list():
            issue_var.set("")

        Button(root, text="▼", font=("Arial", 12), command=show_issues_list, bg="#8B7355", fg="white", bd=0).place(
            x=form_x + 710, y=form_y + 140, width=40, height=40
        )
        # Дата возврата
        import datetime

        today = datetime.date.today()
        date_return_var = StringVar()
        date_return_var.set(today.strftime("%d.%m.%Y"))
        date_return_entry = Entry(
            root, textvariable=date_return_var, font=("Arial", 14), bg="#D2B28F", fg="#2A1E17", bd=2, relief="flat"
        )
        date_return_entry.place(x=form_x + 100, y=form_y + 210, width=600, height=40)
        # Кнопка Готово
        Button(
            root,
            text="Готово",
            font=("Arial", 14),
            bg="#E5D3C6",
            fg="#2A1E17",
            bd=2,
            relief="solid",
            command=lambda: save_return(user_fio, reader_var.get(), issue_var.get(), date_return_var.get()),
        ).place(x=form_x + 250, y=form_y + 280, width=200, height=60)
        # Кнопка Назад
        Button(
            root,
            text="Назад",
            command=lambda: create_staff_home(user_fio),
            bg="#8B7355",
            fg="white",
            activebackground="#6B5A4D",
            activeforeground="white",
        ).place(x=20, y=20)
        # Кнопка Выход
        Button(
            root,
            text="Выход",
            command=logout,
            bg="#8B7355",
            fg="white",
            activebackground="#6B5A4D",
            activeforeground="white",
        ).place(x=width - 120, y=20)


    def save_return(user_fio, fio, issue_str, date_return):
        """Сохраняет возврат в БД и возвращает на главную"""
        try:
            # Получаем id выдачи
            m = re.match(r"ID (\d+)", issue_str)
            if not m:
                messagebox.showerror("Ошибка", "Выберите выдачу!")
                return
            issue_id = int(m.group(1))
            conn = psycopg2.connect(**DB_CONFIG)
            cursor = conn.cursor()
            # Обновляем дату возврата и статус
            cursor.execute(
                """
                UPDATE "Выдача"
                SET "Дата возврата" = %s, "Создана" = TRUE
                WHERE "id" = %s
            """,
                (datetime.datetime.strptime(date_return, "%d.%m.%Y").date(), issue_id),
            )
            # Возвращаем книги на склад (ничего делать не нужно, т.к. количество считается динамически)
            conn.commit()
            conn.close()
            messagebox.showinfo("Успех", "Возврат успешно оформлен!")
            create_staff_home(user_fio)
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при оформлении возврата: {str(e)}")


    def create_book_list_window(user_fio, page=0):
        """Окно просмотра списка книг с карточками и кнопками"""
        clear_window()
        canvas.config(bg="#D5C5AE")
        header_height = 60
        canvas.create_rectangle(0, 0, width, header_height, fill="#54101D", stipple="gray50", outline="")
        Label(root, text=user_fio, font=("Arial", 16), bg="#54101D", fg="white").place(
            x=width // 2, y=header_height // 2, anchor="center"
        )
        initials = "".join([name[0] for name in user_fio.split()[:2]])
        circle_radius = 25
        circle_x = width - 40
        circle_y = header_height // 2
        initials_btn = Button(
            root,
            text=initials,
            font=("Arial", 12),
            bg="white",
            fg="#54101D",
            bd=0,
            command=lambda: create_staff_account(user_fio),
        )
        initials_btn.place(
            x=circle_x - circle_radius, y=circle_y - circle_radius, width=circle_radius * 2, height=circle_radius * 2
        )
        Label(root, text="Список книг", font=("Arial", 16), bg="#D5C5AE", fg="#54101D").place(
            x=width // 2, y=header_height + 30, anchor="center"
        )
        # Получаем список книг
        try:
            conn = psycopg2.connect(**DB_CONFIG)
            cursor = conn.cursor()
            cursor.execute(
                """
                SELECT k."id", k."Название", k."Автор", k."Ценность",
                    COALESCE((SELECT SUM(sp."Количество") FROM "Состав поставки" sp WHERE sp."Книга" = k."id"), 0) as total,
                    COALESCE((SELECT SUM(sv."Количество") FROM "Состав выдачи" sv JOIN "Выдача" v ON sv."Выдача" = v."id" WHERE sv."Книга" = k."id" AND v."Создана" = FALSE), 0) as issued
                FROM "Книги" k
                ORDER BY k."Название" ASC
            """
            )
            all_books = cursor.fetchall()
            conn.close()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка получения списка книг: {str(e)}")
            all_books = []
        books_per_page = 4
        total_pages = max(1, (len(all_books) + books_per_page - 1) // books_per_page)
        current_page = page % total_pages
        start_idx = current_page * books_per_page
        end_idx = min(start_idx + books_per_page, len(all_books))
        books_to_show = all_books[start_idx:end_idx]
        # Карточки книг
        card_width = 250
        card_height = 250
        gap = 30
        start_x = (width - (card_width + gap) * len(books_to_show) + gap) // 2
        y = header_height + 80
        for i, (book_id, title, author, value, total, issued) in enumerate(books_to_show):
            x = start_x + i * (card_width + gap)
            frame = Frame(root, bg="#F8F2EB", highlightbackground="#E5D3C6", highlightthickness=4)
            frame.place(x=x, y=y, width=card_width, height=card_height)
            Label(frame, text=title, font=("Arial", 14), bg="#F8F2EB", fg="#2A1E17", wraplength=220, justify="center").pack(
                pady=(10, 0)
            )
            Label(
                frame,
                text=f"Автор: {author}",
                font=("Arial", 12),
                bg="#F8F2EB",
                fg="#2A1E17",
                wraplength=220,
                justify="center",
            ).pack()
            Label(frame, text=f"Количество: {total - issued} шт", font=("Arial", 12), bg="#F8F2EB", fg="#2A1E17").pack()
            Label(frame, text=f"Ценность за ед: {value}", font=("Arial", 12), bg="#F8F2EB", fg="#2A1E17").pack()
            Label(frame, text=f"Всего: {total}", font=("Arial", 12), bg="#F8F2EB", fg="#2A1E17").pack()
            # Кнопки
            btn_frame = Frame(frame, bg="#F8F2EB")
            btn_frame.pack(pady=5)
            Button(
                btn_frame,
                text="Редактировать",
                font=("Arial", 10),
                command=lambda bid=book_id: edit_book(bid, user_fio),
                bg="#8B7355",
                fg="white",
                bd=0,
            ).pack(side=LEFT, padx=2)
            Button(
                btn_frame,
                text="Удалить",
                font=("Arial", 10),
                command=lambda bid=book_id: confirm_delete_book(bid, user_fio, page, "name", "asc"),
                bg="#8B7355",
                fg="white",
                bd=0,
            ).pack(side=LEFT, padx=2)
            Button(
                frame,
                text="Выданы",
                font=("Arial", 10),
                command=lambda bid=book_id: show_issued_books(bid),
                bg="#E5D3C6",
                fg="#2A1E17",
                bd=0,
            ).pack(pady=2)
        # Пагинация
        nav_frame = Frame(root, bg="#D5C5AE")
        nav_frame.place(x=width // 2 - 75, y=y + card_height + 20, width=200, height=40)
        Button(
            nav_frame,
            text="Назад",
            command=lambda: create_book_list_window(user_fio, current_page - 1),
            bg="#8B7355",
            fg="white",
        ).pack(side=LEFT, padx=10)
        Button(
            nav_frame,
            text="Далее",
            command=lambda: create_book_list_window(user_fio, current_page + 1),
            bg="#8B7355",
            fg="white",
        ).pack(side=LEFT, padx=10)
        # Кнопка Назад
        Button(root, text="Назад", command=lambda: create_staff_home(user_fio), bg="#8B7355", fg="white").place(x=20, y=20)
        # Кнопка Выход
        Button(root, text="Выход", command=logout, bg="#8B7355", fg="white").place(x=width - 120, y=20)


    def show_issued_books(book_id):
        """Показывает окно с выданными экземплярами книги (заглушка)"""
        messagebox.showinfo("Выданы", "Здесь будет список выданных экземпляров книги.")


    def create_readers_with_books_window(user_fio, page=0):
        """Окно со списком читателей, у которых есть книги, с пагинацией"""
        clear_window()
        canvas.config(bg="#D5C5AE")
        header_height = 60
        canvas.create_rectangle(0, 0, width, header_height, fill="#54101D", stipple="gray50", outline="")
        Label(root, text=user_fio, font=("Arial", 16), bg="#54101D", fg="white").place(
            x=width // 2, y=header_height // 2, anchor="center"
        )
        initials = "".join([name[0] for name in user_fio.split()[:2]])
        circle_radius = 25
        circle_x = width - 40
        circle_y = header_height // 2
        initials_btn = Button(
            root,
            text=initials,
            font=("Arial", 12),
            bg="white",
            fg="#54101D",
            bd=0,
            command=lambda: create_staff_account(user_fio),
        )
        initials_btn.place(
            x=circle_x - circle_radius, y=circle_y - circle_radius, width=circle_radius * 2, height=circle_radius * 2
        )
        Label(root, text="Читатели", font=("Arial", 16), bg="#D5C5AE", fg="#54101D").place(
            x=width // 2, y=header_height + 30, anchor="center"
        )
        # Получаем всех читателей
        try:
            conn = psycopg2.connect(**DB_CONFIG)
            cursor = conn.cursor()
            cursor.execute('SELECT "id", "ФИО" FROM "Читатели" ORDER BY "ФИО" ASC')
            readers = cursor.fetchall()
            conn.close()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка получения читателей: {str(e)}")
            readers = []
        # Пагинация
        readers_per_page = 4
        total_pages = max(1, (len(readers) + readers_per_page - 1) // readers_per_page)
        current_page = page % total_pages
        start_idx = current_page * readers_per_page
        end_idx = min(start_idx + readers_per_page, len(readers))
        readers_to_show = readers[start_idx:end_idx]
        # Кнопки-читатели
        form_width = 800
        form_height = 400
        form_x = (width - form_width) // 2
        form_y = header_height + 50
        round_rectangle(
            form_x, form_y, form_x + form_width, form_y + form_height, radius=50, fill="#D2B28F", stipple="gray50"
        )
        y0 = form_y + 40
        for i, (reader_id, fio) in enumerate(readers_to_show):
            Button(
                root,
                text=fio,
                font=("Arial", 14),
                bg="#C3A580",
                fg="#2A1E17",
                bd=0,
                command=lambda rid=reader_id, fio=fio: create_reader_books_window(user_fio, rid, fio),
            ).place(x=form_x + 50, y=y0 + i * 60, width=700, height=40)
        # Пагинация
        nav_frame = Frame(root, bg="#D5C5AE")
        nav_frame.place(x=form_x + 400 - 100, y=form_y + form_height - 70, width=200, height=50)
        Button(
            nav_frame,
            text="Назад",
            font=("Arial", 14),
            bg="#8B7355",
            fg="white",
            command=lambda: create_readers_with_books_window(user_fio, (current_page - 1) % total_pages),
        ).pack(side=LEFT, padx=10)
        Button(
            nav_frame,
            text="Далее",
            font=("Arial", 14),
            bg="#E5D3C6",
            fg="#2A1E17",
            command=lambda: create_readers_with_books_window(user_fio, (current_page + 1) % total_pages),
        ).pack(side=LEFT, padx=10)
        Button(root, text="Назад", command=lambda: create_staff_home(user_fio), bg="#8B7355", fg="white").place(x=20, y=20)
        Button(root, text="Выход", command=logout, bg="#8B7355", fg="white").place(x=width - 120, y=20)


    def create_reader_books_window(user_fio, reader_id, fio):
        """Окно с читательским билетом и списком книг читателя"""
        clear_window()
        canvas.config(bg="#D5C5AE")
        header_height = 60
        canvas.create_rectangle(0, 0, width, header_height, fill="#54101D", stipple="gray50", outline="")
        Label(root, text=user_fio, font=("Arial", 16), bg="#54101D", fg="white").place(
            x=width // 2, y=header_height // 2, anchor="center"
        )
        initials = "".join([name[0] for name in user_fio.split()[:2]])
        circle_radius = 25
        circle_x = width - 40
        circle_y = header_height // 2
        initials_btn = Button(
            root,
            text=initials,
            font=("Arial", 12),
            bg="white",
            fg="#54101D",
            bd=0,
            command=lambda: create_staff_account(user_fio),
        )
        initials_btn.place(
            x=circle_x - circle_radius, y=circle_y - circle_radius, width=circle_radius * 2, height=circle_radius * 2
        )
        Label(root, text=f"{fio}", font=("Arial", 16), bg="#D5C5AE", fg="#54101D").place(
            x=width // 2, y=header_height + 30, anchor="center"
        )
        # Получаем id читательского билета
        Label(root, text=f"Читательский билет: {reader_id}", font=("Arial", 14), bg="#C3A580", fg="#2A1E17").place(
            x=width // 2 - 200, y=header_height + 70, width=400, height=40
        )
        # Получаем список не возвращённых книг
        try:
            conn = psycopg2.connect(**DB_CONFIG)
            cursor = conn.cursor()
            cursor.execute(
                """
                SELECT k."Название", v."Дата возврата", v."Дата выдачи", sv."Количество"
                FROM "Выдача" v
                JOIN "Состав выдачи" sv ON v."id" = sv."Выдача"
                JOIN "Книги" k ON sv."Книга" = k."id"
                WHERE v."Читатель" = %s AND v."Создана" = FALSE
                ORDER BY v."Дата возврата" ASC
            """,
                (reader_id,),
            )
            books = cursor.fetchall()
            conn.close()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка получения книг: {str(e)}")
            books = []
        # Таблица книг
        table_x = width // 2 - 350
        table_y = header_height + 130
        col_widths = [220, 150, 150, 120]
        headers = ["Название", "Дата возврата", "Дата взятия", "Количество"]
        for j, h in enumerate(headers):
            Label(root, text=h, font=("Arial", 12, "bold"), bg="#C3A580", fg="#2A1E17").place(
                x=table_x + sum(col_widths[:j]), y=table_y, width=col_widths[j], height=40
            )
        for i, (title, date_return, date_issue, quantity) in enumerate(books):
            Label(root, text=title, font=("Arial", 12), bg="#D2B28F", fg="#2A1E17", wraplength=200, justify="center").place(
                x=table_x, y=table_y + 50 + i * 50, width=col_widths[0], height=40
            )
            Label(root, text=date_return.strftime("%d.%m.%Y"), font=("Arial", 12), bg="#D2B28F", fg="#2A1E17").place(
                x=table_x + col_widths[0], y=table_y + 50 + i * 50, width=col_widths[1], height=40
            )
            Label(root, text=date_issue.strftime("%d.%m.%Y"), font=("Arial", 12), bg="#D2B28F", fg="#2A1E17").place(
                x=table_x + col_widths[0] + col_widths[1], y=table_y + 50 + i * 50, width=col_widths[2], height=40
            )
            Label(root, text=str(quantity), font=("Arial", 12), bg="#D2B28F", fg="#2A1E17").place(
                x=table_x + col_widths[0] + col_widths[1] + col_widths[2],
                y=table_y + 50 + i * 50,
                width=col_widths[3],
                height=40,
            )
        Button(
            root, text="Назад", command=lambda: create_readers_with_books_window(user_fio), bg="#8B7355", fg="white"
        ).place(x=width - 200, y=header_height + 30, width=120, height=40)
        Button(root, text="Далее", font=("Arial", 14), bg="#E5D3C6", fg="#2A1E17", bd=2, relief="solid").place(
            x=width - 200, y=table_y + 50 + max(len(books), 1) * 50, width=150, height=50
        )
        Button(root, text="Выход", command=logout, bg="#8B7355", fg="white").place(x=width - 120, y=20)


    def create_export_window(user_fio):
        """Окно для выбора типа данных для экспорта"""
        if openpyxl is None:
            messagebox.showerror("Ошибка", "Для экспорта в Excel требуется установить openpyxl: pip install openpyxl")
            return
        export_win = Toplevel(root)
        export_win.title("Выгрузить данные")
        export_win.geometry("400x300")
        export_win.config(bg="#E5D3C6")
        Label(export_win, text="Что выгрузить?", font=("Arial", 14), bg="#E5D3C6", fg="#54101D").pack(pady=20)
        Button(export_win, text="Все книги", font=("Arial", 12), command=lambda: export_books(export_win)).pack(
            pady=10, fill="x", padx=40
        )
        Button(export_win, text="Все читатели", font=("Arial", 12), command=lambda: export_readers(export_win)).pack(
            pady=10, fill="x", padx=40
        )
        Button(export_win, text="Все выдачи", font=("Arial", 12), command=lambda: export_issues(export_win)).pack(
            pady=10, fill="x", padx=40
        )
        Button(export_win, text="Закрыть", font=("Arial", 12), command=export_win.destroy).pack(pady=20)


    def export_books(win):
        try:
            conn = psycopg2.connect(**DB_CONFIG)
            cursor = conn.cursor()
            cursor.execute('SELECT "Название", "Автор", "Ценность", "Количество" FROM "Книги" ORDER BY "Название" ASC')
            books = cursor.fetchall()
            conn.close()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Книги"
            ws.append(["Название", "Автор", "Ценность", "Количество"])
            for cell in ws[1]:
                cell.font = Font(bold=True)
            for row in books:
                ws.append(row)
            filename = f"Книги_{get_export_timestamp()}.xlsx"
            wb.save(filename)
            messagebox.showinfo("Успех", f"Данные о книгах выгружены в файл {filename}")
            win.destroy()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка экспорта: {str(e)}")


    def export_readers(win):
        try:
            conn = psycopg2.connect(**DB_CONFIG)
            cursor = conn.cursor()
            cursor.execute('SELECT "ФИО", "Адрес", "Номер телефона" FROM "Читатели" ORDER BY "ФИО" ASC')
            readers = cursor.fetchall()
            conn.close()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Читатели"
            ws.append(["ФИО", "Адрес", "Номер телефона"])
            for cell in ws[1]:
                cell.font = Font(bold=True)
            for row in readers:
                ws.append(row)
            filename = f"Читатели_{get_export_timestamp()}.xlsx"
            wb.save(filename)
            messagebox.showinfo("Успех", f"Данные о читателях выгружены в файл {filename}")
            win.destroy()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка экспорта: {str(e)}")


    def export_issues(win):
        try:
            conn = psycopg2.connect(**DB_CONFIG)
            cursor = conn.cursor()
            cursor.execute(
                """
                SELECT v."id", s."ФИО", st."ФИО", k."Название", sv."Количество", v."Дата выдачи", v."Дата возврата", v."Создана"
                FROM "Выдача" v
                JOIN "Читатели" s ON v."Читатель" = s."id"
                JOIN "Сотрудник" st ON v."Сотрудник" = st."id"
                JOIN "Состав выдачи" sv ON v."id" = sv."Выдача"
                JOIN "Книги" k ON sv."Книга" = k."id"
                ORDER BY v."Дата выдачи" DESC
            """
            )
            issues = cursor.fetchall()
            conn.close()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Выдачи"
            ws.append(
                ["ID выдачи", "Читатель", "Сотрудник", "Книга", "Кол-во", "Дата выдачи", "Дата возврата", "Возвращено"]
            )
            for cell in ws[1]:
                cell.font = Font(bold=True)
            for row in issues:
                ws.append(
                    [
                        row[0],
                        row[1],
                        row[2],
                        row[3],
                        row[4],
                        row[5].strftime("%d.%m.%Y"),
                        row[6].strftime("%d.%m.%Y"),
                        "Да" if row[7] else "Нет",
                    ]
                )
            filename = f"Выдачи_{get_export_timestamp()}.xlsx"
            wb.save(filename)
            messagebox.showinfo("Успех", f"Данные о выдачах выгружены в файл {filename}")
            win.destroy()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка экспорта: {str(e)}")


    def get_export_timestamp():
        import datetime

        return datetime.datetime.now().strftime("%Y%m%d_%H%M%S")


    root.mainloop()
